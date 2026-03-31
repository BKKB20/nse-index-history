import json, csv, re
from pathlib import Path
from datetime import datetime, date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE        = Path(r"C:\Users\Bhavya Khaitan\BK\Stocks\index_reconstructor")
PDF_JSON    = BASE / "all_circulars_parsed_v3.json"
HTM_JSON    = BASE / "htm_circulars_parsed_v2.json"
MEMBERS_DIR = BASE / "current_members"
OUTPUT_DIR  = BASE / "output_v2"
OUTPUT_DIR.mkdir(exist_ok=True)
OUT_FILE    = OUTPUT_DIR / "NSE_INDEX_HISTORY_FINAL_V3.xlsx"

INDEX_CSV = {
    "NIFTY 50":                 "NIFTY_50.csv",
    "NIFTY NEXT 50":            "NIFTY_NEXT_50.csv",
    "NIFTY 100":                "NIFTY_100.csv",
    "NIFTY 500":                "NIFTY_500.csv",
    "NIFTY BANK":               "NIFTY_BANK.csv",
    "NIFTY IT":                 "NIFTY_IT.csv",
    "NIFTY FMCG":               "NIFTY_FMCG.csv",
    "NIFTY AUTO":               "NIFTY_AUTO.csv",
    "NIFTY PHARMA":             "NIFTY_PHARMA.csv",
    "NIFTY ENERGY":             "NIFTY_ENERGY.csv",
    "NIFTY METAL":              "NIFTY_METAL.csv",
    "NIFTY MIDCAP 50":          "NIFTY_MIDCAP_50.csv",
    "NIFTY MIDCAP 150":         "NIFTY_MIDCAP_150.csv",
    "NIFTY SMALLCAP 250":       "NIFTY_SMALLCAP_250.csv",
    "NIFTY FINANCIAL SERVICES": "NIFTY_FINANCIAL_SERVICES.csv",
    "NIFTY PSU BANK":           "NIFTY_PSU_BANK.csv",
    "NIFTY PRIVATE BANK":       "NIFTY_PRIVATE_BANK.csv",
    "NIFTY COMMODITIES":        "NIFTY_COMMODITIES.csv",
    "NIFTY REALTY":             "NIFTY_REALTY.csv",
    "NIFTY INFRASTRUCTURE":     "NIFTY_INFRASTRUCTURE.csv",
    "NIFTY MEDIA":              "NIFTY_MEDIA.csv",
    "NIFTY CPSE":               "NIFTY_CPSE.csv",
    "NIFTY MNC":                "NIFTY_MNC.csv",
    "NIFTY INDIA CONSUMPTION":  "NIFTY_INDIA_CONSUMPTION.csv",
    "NIFTY ALPHA 50":           "NIFTY_ALPHA_50.csv",
    "NIFTY LOW VOLATILITY 50":  "NIFTY_LOW_VOLATILITY_50.csv",
    "NIFTY MANUFACTURING":      "NIFTY_MANUFACTURING.csv",
    "NIFTY INDIA DEFENCE":      "NIFTY_INDIA_DEFENCE.csv",
}

INDEX_SIZE = {
    "NIFTY 50":50,"NIFTY NEXT 50":50,"NIFTY 100":100,"NIFTY 500":500,
    "NIFTY BANK":14,"NIFTY IT":10,"NIFTY FMCG":15,"NIFTY AUTO":15,
    "NIFTY PHARMA":20,"NIFTY ENERGY":40,"NIFTY METAL":15,
    "NIFTY MIDCAP 50":50,"NIFTY MIDCAP 150":150,"NIFTY SMALLCAP 250":250,
    "NIFTY FINANCIAL SERVICES":20,"NIFTY PSU BANK":12,"NIFTY PRIVATE BANK":10,
    "NIFTY COMMODITIES":30,"NIFTY REALTY":10,"NIFTY INFRASTRUCTURE":30,
    "NIFTY MEDIA":10,"NIFTY CPSE":11,"NIFTY MNC":30,
    "NIFTY INDIA CONSUMPTION":30,"NIFTY ALPHA 50":50,
    "NIFTY LOW VOLATILITY 50":50,"NIFTY MANUFACTURING":74,
    "NIFTY INDIA DEFENCE":18,
}

ALIAS = {
    "HDFC":"HDFCBANK","MINDTREE":"LTIM","LTINFOTECH":"LTIM",
    "PVR":"PVRINOX","INOXLEISUR":"PVRINOX","INFRATEL":"BHARTIARTL",
    "ORIENTBANK":"PNB","CORPBANK":"UNIONBANK","SYNDIBANK":"CANARABANK",
    "ALBK":"INDIANB","ALLBANK":"INDIANB","ANDHRBANK":"UNIONBANK",
    "SESAGOA":"VEDL","CAIRN":"VEDL","NIITTECH":"COFORGE",
    "HDFCSTD":"HDFCLIFE","RECLTD":"REC","ACCLTD":"ACC","JSWISPAT":"JSWSTEEL",
    "RELPETRO":"RELIANCE","SATYAMCOMP":"TECHM","IGATE":"TECHM",
    "SUBEXAZURELTD":"SUBEX","MCLEODRUSS":"MCLEODRUSS",
    "UTIBANKLTD":"AXISBANK","UTIBANK":"AXISBANK",
}

# Tickers that are clearly garbage from concatenation — skip them
GARBAGE = {
    "DLFLTDDLF","PFIZERLTDPFIZER","SUBEXLTDSUBEX","AMTEKAUTOLTD",
    "WOCKHARDTLTD","PENINSULALANDLTD","KESORAMIND","IDBIBANKLTD",
    "JSWSTEELLTD","UNIPHOS","CUMMINSINDIALTD","CORPORATIONBANK",
    "TECHMAHINDRALTD","NIITLTD","IDFCLTD","HTMEDIALTD","MERCATORLINESLTD",
    "GREAVESCOTTONLTD","INDIAGLYCOLSLTD","KDLBIOTECHLTD","BSLLTD",
    "SUZLONENERGYLTD","SIEMENSLTD","TATATEALTD","BANKOFINDIA",
    "RELPETRO","DABURINDIALTD","NTPCLTD","MASTEKLTD","UNITECHLIMITED",
    "IDEACELLULARLTD","CAIRNINDIALTD","HINDUJATMTLTD","SUBEXAZURELTD",
    "AMTEKINDIALTD","JAINSTUDIOSLTD","CINEVISTAASLTD","GTNINDUSTRIESLTD",
    "BOMDYEING","JSWSTEELLTD","TATACHEMICALSLTD","UNITEDSPIRITSLTD",
    "INDIACEMENTSLTD","MPHASISLTD","ELECTRONICS",
    "PHARMACEUTICALS","GAS","METALS","MISCELLANEOUS","CONSTRUCTION",
    "AUTOANCILLARIES","TEXTILES","FERTILISERS","PACKAGING","PETROCHEMICALS",
    "FINANCE","COMPUTERS","SOFTWARE","RETAIL","BREW","DISTILLERIES",
    "SHIPPING","LARSEN","TOUBROLTD","BIOCONLTD","LUPINLTD","MONNETISPATLTD",
    "NDTVLTD","PETRONETLNGLTD","RADICOKHAITANLTD","TRENTLTD","VSTINDUSTRIESLTD",
    "HMTLTD","ANDHRABANK","BONGREFIN","STERLINBIO",
    "INDONATIONALLTD","CMCLTD","IBPCOLTD","MICROINKSLTD","MICROINKS",
    "TATACHEMICALSLTD","GMRINFRA","POLARIS","POWER","IBPCOLTD",
    "TULIPTELECOMLTD","AMTEKAUTOLTD","RELCAPITAL",
}

# Fix known wrong tickers from HTM
FIX_MAP = {
    "TECHMAHINDRALTD":  "TECHM",
    "ULTRACEMCO":       "ULTRACEMCO",
    "UTIBANKLTD":       "AXISBANK",
    "DABURINDIALTD":    "DABUR",
    "NTPCLTD":          "NTPC",
    "MASTEKLTD":        "MASTEK",
    "UNITECHLIMITED":   "UNITECH",
    "IDEACELLULARLTD":  "IDEA",
    "CAIRNINDIALTD":    "CAIRN",
    "SUZLONENERGYLTD":  "SUZLON",
    "SIEMENSLTD":       "SIEMENS",
    "MPHASISLTD":       "MPHASIS",
    "SUBEXAZURELTD":    "SUBEX",
    "IDBIBANKLTD":      "IDBI",
    "JSWSTEELLTD":      "JSWSTEEL",
    "CUMMINSINDIALTD":  "CUMMINSIND",
    "WOCKHARDTLTD":     "WOCKPHARMA",
    "AMTEKAUTOLTD":     "AMTEKAUTO",
    "IDFCLTD":          "IDFC",
    "BANKOFINDIA":      "BANKINDIA",
    "SYNDICATEBANK":    "SYNDIBANK",
    "TATACHEMICALSLTD": "TATACHEM",
    "UNITEDSPIRITSLTD": "MCDOWELL-N",
    "INDIACEMENTSLTD":  "INDIACEM",
    "BIOCONLTD":        "BIOCON",
    "LUPINLTD":         "LUPIN",
    "TRENTLTD":         "TRENT",
    "RADICOKHAITANLTD": "RADICO",
    "VSTINDUSTRIESLTD": "VSTIND",
    "RELPETRO":         "RELIANCE",
    "HMTLTD":           "HMT",
    "CMCLTD":           "CMC",
    "IBPCOLTD":         "IBP",
    "GMRINFRA":         "GMRINFRA",
    "HINDUJATMTLTD":    "HINDUJATMT",
    "AMTEKINDIALTD":    "AMTEKAUTO",
    "CORPORATIONBANK":  "CORPBANK",
    "KOTAKBANK":        "KOTAKBANK",
    "DLFLTDDLF":        None,  # garbage
    "PFIZERLTDPFIZER":  "PFIZER",
    "SUBEXLTDSUBEX":    None,
    "HDIL":             "HDIL",
    "INDIANBANK":       "INDIANB",
    "ALLCARGO":         "ALLCARGO",
    "BRFL":             "BRFL",
    "MCLEODRUSS":       "MCLEODRUSS",
    "NIITLTD":          "NIITLTD",
}

def clean_ticker(sym):
    """Clean and validate a ticker symbol."""
    s = re.sub(r'[^A-Z0-9&\-]', '', str(sym).strip().upper())
    # Apply fix map first
    if s in FIX_MAP:
        s = FIX_MAP[s]
        if s is None:
            return None
    # Apply alias
    s = ALIAS.get(s, s)
    # Reject garbage
    if s in GARBAGE:
        return None
    # Must look like a valid ticker
    if not re.match(r'^[A-Z0-9][A-Z0-9&\-]{1,19}$', s):
        return None
    if len(s) < 2 or len(s) > 20:
        return None
    # Reject if it looks like a company name (too long or has lowercase)
    if len(s) > 15:
        return None
    return s

def load_members(index_name):
    csv_file = INDEX_CSV.get(index_name)
    if not csv_file:
        return set()
    path = MEMBERS_DIR / csv_file
    if not path.exists():
        return set()
    members = set()
    with open(path, encoding="utf-8-sig", errors="ignore") as f:
        for row in csv.DictReader(f):
            for col in row:
                if col.strip().upper() == 'SYMBOL':
                    v = str(row[col]).strip()
                    t = clean_ticker(v)
                    if t:
                        members.add(t)
                    break
    return members

def load_all_circulars():
    """Load and merge PDF v3 + HTM v2 circulars, clean tickers."""
    records = []
    for json_file in [PDF_JSON, HTM_JSON]:
        if not json_file.exists():
            continue
        with open(json_file, encoding="utf-8") as f:
            data = json.load(f)
        for rec in data:
            try:
                d = datetime.strptime(rec["date"][:10], "%Y-%m-%d").date()
            except:
                continue
            inc = [t for t in (clean_ticker(s) for s in rec.get("inclusions",[])) if t]
            exc = [t for t in (clean_ticker(s) for s in rec.get("exclusions",[])) if t]
            if not inc and not exc:
                continue
            records.append({
                "date":       d,
                "index":      rec["index"],
                "inclusions": inc,
                "exclusions": exc,
                "ref":        rec.get("ref",""),
                "source":     rec.get("source",""),
            })
    return sorted(records, key=lambda x: x["date"])

def reconstruct_index(index_name, all_circulars):
    current = load_members(index_name)
    if not current:
        return {}, []

    relevant = [c for c in all_circulars if c["index"] == index_name]
    relevant_desc = list(reversed(relevant))

    TODAY = date.today()
    state = set(current)
    snapshots = {TODAY: frozenset(state)}
    anomalies = []
    expected = INDEX_SIZE.get(index_name)

    for circ in relevant_desc:
        d   = circ["date"]
        inc = circ["inclusions"]
        exc = circ["exclusions"]
        ref = circ["ref"]

        new_state = set(state)
        skipped_inc = []

        for s in inc:
            if s in new_state:
                new_state.discard(s)
            else:
                skipped_inc.append(s)

        for s in exc:
            new_state.add(s)

        if skipped_inc:
            anomalies.append({
                "date": str(d), "ref": ref,
                "issue": f"INC_NOT_IN_STATE: {skipped_inc}"
            })

        state = new_state
        snapshots[d] = frozenset(state)

    return snapshots, anomalies

def get_on_date(snapshots, query_date):
    if isinstance(query_date, str):
        query_date = datetime.strptime(query_date, "%Y-%m-%d").date()
    result = None
    for d in sorted(snapshots):
        if d <= query_date:
            result = snapshots[d]
    return result or frozenset()

# ── MAIN ──────────────────────────────────────────────────────────────────────
print("Loading and merging circulars...")
all_circulars = load_all_circulars()
print(f"Total clean circular records: {len(all_circulars)}")

from collections import Counter
idx_counts = Counter(c["index"] for c in all_circulars)
print("\nRecords per index:")
for k,v in sorted(idx_counts.items()):
    print(f"  {k:35s} {v}")

print("\nReconstructing...")
all_snapshots = {}
summary_rows  = []

for idx in INDEX_CSV:
    snaps, anom = reconstruct_index(idx, all_circulars)
    all_snapshots[idx] = snaps
    today_m  = get_on_date(snaps, date.today())
    expected = INDEX_SIZE.get(idx, "?")
    status   = "OK" if (expected=="?" or len(today_m)==expected) else "MISMATCH"
    print(f"  {idx:35s} {len(snaps):3} snaps  "
          f"{len(today_m):4} today  exp={expected}  {status}  anom={len(anom)}")
    summary_rows.append({
        "Index":idx,"Snapshots":len(snaps),
        "Today_Count":len(today_m),"Expected":expected,
        "Status":status,"Anomalies":len(anom),
    })

# ── Build sheets ──────────────────────────────────────────────────────────────
print("\nBuilding Excel sheets...")
today = date.today()

# Sheet 1: FLAT_LIST
flat_rows = []
for idx, snaps in all_snapshots.items():
    for d in sorted(snaps):
        for sym in sorted(snaps[d]):
            flat_rows.append({"Effective_Date":str(d),"Index":idx,"Symbol":sym})
df_flat = pd.DataFrame(flat_rows)

# Sheet 2: CONSTITUENTS_BY_DATE
all_dates = sorted(set(d for s in all_snapshots.values() for d in s))
date_rows = []
for d in all_dates:
    row = {"Effective_Date":str(d)}
    for idx in INDEX_CSV:
        m = get_on_date(all_snapshots[idx], d)
        row[idx] = ", ".join(sorted(m)) if m else ""
    date_rows.append(row)
df_dates = pd.DataFrame(date_rows)

# Sheet 3: WIDE_FORMAT with TOTAL row
wide_frames = []
for idx in INDEX_CSV:
    snaps = all_snapshots[idx]
    dates = sorted(snaps.keys())
    all_syms = sorted(set(s for fs in snaps.values() for s in fs))
    if not all_syms: continue
    # TOTAL row first
    sum_row = {"Index":idx,"Symbol":"TOTAL"}
    for d in dates:
        sum_row[str(d)] = sum(1 if sym in snaps[d] else 0 for sym in all_syms)
    rows = [sum_row]
    for sym in all_syms:
        row = {"Index":idx,"Symbol":sym}
        for d in dates:
            row[str(d)] = 1 if sym in snaps[d] else 0
        rows.append(row)
    wide_frames.append(pd.DataFrame(rows))
df_wide = pd.concat(wide_frames, ignore_index=True) if wide_frames else pd.DataFrame()

# Sheet 4: TODAY
today_data = {}
max_len = 0
for idx in INDEX_CSV:
    m = sorted(get_on_date(all_snapshots[idx], today))
    today_data[idx] = m
    max_len = max(max_len, len(m))
today_rows = []
for i in range(max_len):
    row = {}
    for idx in INDEX_CSV:
        lst = today_data[idx]
        row[idx] = lst[i] if i < len(lst) else ""
    today_rows.append(row)
df_today = pd.DataFrame(today_rows)

# Sheet 5: SUMMARY
df_summary = pd.DataFrame(summary_rows)

# Sheet 6: INDEX_INFO
INDEX_INFO = [
    ("NIFTY 50","April 1996","Broad Market","NSE Indices Limited"),
    ("NIFTY NEXT 50","January 1997","Broad Market","NSE Indices Limited"),
    ("NIFTY 100","January 2003","Broad Market","NSE Indices Limited"),
    ("NIFTY 500","January 1995","Broad Market","NSE Indices Limited"),
    ("NIFTY BANK","September 2003","Sectoral - Banking","NSE Indices Limited"),
    ("NIFTY IT","January 1996","Sectoral - Technology","NSE Indices Limited"),
    ("NIFTY FMCG","January 1996","Sectoral - FMCG","NSE Indices Limited"),
    ("NIFTY AUTO","January 2004","Sectoral - Auto","NSE Indices Limited"),
    ("NIFTY PHARMA","January 2001","Sectoral - Pharma","NSE Indices Limited"),
    ("NIFTY ENERGY","January 2001","Sectoral - Energy","NSE Indices Limited"),
    ("NIFTY METAL","January 2004","Sectoral - Metal","NSE Indices Limited"),
    ("NIFTY MIDCAP 50","July 2004","Midcap","NSE Indices Limited"),
    ("NIFTY MIDCAP 150","April 2016","Midcap","NSE Indices Limited"),
    ("NIFTY SMALLCAP 250","April 2016","Smallcap","NSE Indices Limited"),
    ("NIFTY FINANCIAL SERVICES","January 2004","Sectoral - Finance","NSE Indices Limited"),
    ("NIFTY PSU BANK","January 2004","Sectoral - Banking","NSE Indices Limited"),
    ("NIFTY PRIVATE BANK","January 2004","Sectoral - Banking","NSE Indices Limited"),
    ("NIFTY COMMODITIES","July 2011","Thematic","NSE Indices Limited"),
    ("NIFTY REALTY","January 2007","Sectoral - Realty","NSE Indices Limited"),
    ("NIFTY INFRASTRUCTURE","January 2004","Thematic","NSE Indices Limited"),
    ("NIFTY MEDIA","July 2005","Sectoral - Media","NSE Indices Limited"),
    ("NIFTY CPSE","January 2009","Thematic - PSU","NSE Indices Limited"),
    ("NIFTY MNC","January 1995","Thematic - MNC","NSE Indices Limited"),
    ("NIFTY INDIA CONSUMPTION","January 2009","Thematic","NSE Indices Limited"),
    ("NIFTY ALPHA 50","April 2012","Strategy - Factor","NSE Indices Limited"),
    ("NIFTY LOW VOLATILITY 50","July 2012","Strategy - Factor","NSE Indices Limited"),
    ("NIFTY MANUFACTURING","May 2019","Thematic","NSE Indices Limited"),
    ("NIFTY INDIA DEFENCE","April 2021","Thematic - Defence","NSE Indices Limited"),
]
df_info = pd.DataFrame(INDEX_INFO,
    columns=["Index Name","Launch Date","Category","Index Manager"])

# ── Write Excel ───────────────────────────────────────────────────────────────
print(f"\nWriting {OUT_FILE.name}...")
with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
    df_info.to_excel(writer,    sheet_name="INDEX_INFO",           index=False)
    df_summary.to_excel(writer, sheet_name="SUMMARY",              index=False)
    df_dates.to_excel(writer,   sheet_name="CONSTITUENTS_BY_DATE", index=False)
    df_flat.to_excel(writer,    sheet_name="FLAT_LIST",            index=False)
    df_wide.to_excel(writer,    sheet_name="WIDE_FORMAT",          index=False)
    df_today.to_excel(writer,   sheet_name="TODAY",                index=False)

# Format
print("Formatting...")
wb  = load_workbook(OUT_FILE)
HDR = PatternFill("solid", fgColor="1F4E79")
HF  = Font(bold=True, color="FFFFFF", size=10)
ALT = PatternFill("solid", fgColor="EBF3FB")
NRM = PatternFill("solid", fgColor="FFFFFF")
SUM = PatternFill("solid", fgColor="FFF2CC")
GRN = PatternFill("solid", fgColor="E2EFDA")
RED = PatternFill("solid", fgColor="FFE0E0")
ONE = PatternFill("solid", fgColor="C6EFCE")

for ws in wb.worksheets:
    for cell in ws[1]:
        cell.fill = HDR
        cell.font = HF
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    for i, row in enumerate(ws.iter_rows(min_row=2), 2):
        is_total = (len(row)>1 and str(row[1].value or "")=="TOTAL")
        fill = SUM if is_total else (ALT if i%2==0 else NRM)
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(ml+2, 55)

# Color WIDE_FORMAT 1s green
ws_wide = wb["WIDE_FORMAT"]
for row in ws_wide.iter_rows(min_row=2):
    if str(row[1].value or "") == "TOTAL":
        continue
    for cell in row[2:]:
        if cell.value == 1:
            cell.fill = ONE

# Color SUMMARY status
ws_sum = wb["SUMMARY"]
for row in ws_sum.iter_rows(min_row=2):
    for cell in row:
        if cell.column_letter == "E":
            cell.fill = GRN if str(cell.value)=="OK" else RED

wb.save(OUT_FILE)
print(f"\nDone: {OUT_FILE}")
print(f"\nSheets: INDEX_INFO | SUMMARY | CONSTITUENTS_BY_DATE | FLAT_LIST | WIDE_FORMAT | TODAY")